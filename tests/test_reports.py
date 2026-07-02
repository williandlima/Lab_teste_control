"""Testes de ponta a ponta do gerador de relatórios (dados reais via repositories)."""
from __future__ import annotations

import zipfile
from dataclasses import replace
from pathlib import Path
from xml.etree import ElementTree as ET

import pytest

from config import BrandingConfig
from database.database import Database
from database.models import (
    Board,
    Evaluation,
    EvaluationResult,
    EventLogEntry,
    MonitoredSample,
    Operator,
    PowerStep,
    TestParameterConfig,
    TestSession,
    TestSessionStatus,
)
from database.repositories import (
    BoardRepository,
    EvaluationRepository,
    EventLogRepository,
    MonitoredSampleRepository,
    OperatorRepository,
    TestParameterConfigRepository,
    TestSessionRepository,
)
from reports.excel_report import generate_excel_report
from reports.pdf_report import generate_pdf_report
from reports.report_data import assemble_report_data
from reports.word_report import generate_word_report


@pytest.fixture()
def branding(tmp_path: Path) -> BrandingConfig:
    return BrandingConfig(
        company_name="Avibras Aeroco",
        logo_path=tmp_path / "no_logo.png",
        color_primary_navy="#0A1F44",
        color_secondary_navy="#14315C",
        color_accent_orange="#FF7A29",
        color_accent_orange_hover="#FF9248",
        color_text_on_navy="#F5F7FA",
        color_pass="#2ECC71",
        color_fail="#E74C3C",
        color_warning="#F1C40F",
    )


@pytest.fixture()
def populated_session_id(tmp_path: Path) -> tuple[Database, int]:
    db = Database(tmp_path / "fct_test.db")
    db.connect()

    operator = OperatorRepository(db).get_or_create("Joao Silva")
    board = BoardRepository(db).get_or_create("BRD-001", "PN-123", "A")
    config = TestParameterConfigRepository(db).save(
        TestParameterConfig(
            id=None,
            board_id=board.id,
            name="Config padrão",
            nominal_voltage=5.0,
            voltage_min=4.5,
            voltage_max=5.5,
            current_max=1.0,
            test_duration_s=2.0,
            power_sequence=[PowerStep(voltage=5.0, current=1.0, duration_s=2.0)],
        )
    )
    session = TestSessionRepository(db).create(
        TestSession(
            id=None,
            board_id=board.id,
            serial_number="SN-001",
            operator_id=operator.id,
            test_parameter_config_id=config.id,
            config_snapshot_json=(
                '{"nominal_voltage": 5.0, "voltage_min": 4.5, "voltage_max": 5.5, '
                '"current_max": 1.0, "test_duration_s": 2.0}'
            ),
            production_order="OP-9",
            observations=None,
            status=TestSessionStatus.RUNNING,
            started_at="2026-06-25 10:00:00",
        )
    )
    TestSessionRepository(db).update_status(
        session.id, TestSessionStatus.COMPLETED, finished_at="2026-06-25 10:05:00"
    )
    MonitoredSampleRepository(db).insert_batch(
        [
            MonitoredSample(
                id=None, test_session_id=session.id, timestamp=f"2026-06-25 10:0{i}:00",
                step_index=0, voltage_measured=5.0 + i * 0.01, current_measured=0.5 + i * 0.001,
            )
            for i in range(5)
        ]
    )
    EventLogRepository(db).add(
        EventLogEntry(
            id=None, test_session_id=session.id, timestamp=None, level="INFO",
            source="state_machine", message="Teste concluído.",
        )
    )
    EvaluationRepository(db).create(
        Evaluation(
            id=None, test_session_id=session.id, operator_id=operator.id,
            result=EvaluationResult.APPROVED, comment="OK",
        )
    )
    return db, session.id


def _data(db: Database, session_id: int):
    return assemble_report_data(
        session_id,
        TestSessionRepository(db),
        BoardRepository(db),
        OperatorRepository(db),
        MonitoredSampleRepository(db),
        EvaluationRepository(db),
        EventLogRepository(db),
    )


def test_assemble_report_data_computes_observed_ranges(populated_session_id) -> None:
    db, session_id = populated_session_id
    data = _data(db, session_id)

    assert data.board.code == "BRD-001"
    assert data.operator.name == "Joao Silva"
    assert len(data.samples) == 5
    assert data.evaluation is not None and data.evaluation.result == EvaluationResult.APPROVED
    assert data.voltage_min_observed == pytest.approx(5.0)
    assert data.voltage_max_observed == pytest.approx(5.04)
    assert data.config_snapshot["nominal_voltage"] == 5.0


def test_generate_excel_report_creates_file(populated_session_id, branding, tmp_path: Path) -> None:
    db, session_id = populated_session_id
    data = _data(db, session_id)
    output_dir = tmp_path / "exports"

    path = generate_excel_report(data, branding, output_dir)

    assert path.exists()
    assert path.suffix == ".xlsx"
    assert path.stat().st_size > 0


def test_generate_excel_report_does_not_distort_the_logo(
    populated_session_id, branding, tmp_path: Path
) -> None:
    """A logo real é landscape (~2.12:1) -- versão anterior forçava 60x60px
    (quadrado), esticando/achatando o traço visivelmente. `openpyxl` recalcula
    Image.width/height a partir do PNG bruto ao recarregar (ignora o extent
    salvo), então a única forma confiável de checar o que FOI SALVO é ler o
    <xdr:ext> do drawing XML dentro do .xlsx diretamente."""
    real_logo = Path(__file__).resolve().parent.parent / "assets" / "branding" / "avibras_aeroco_logo.png"
    db, session_id = populated_session_id
    data = _data(db, session_id)
    output_dir = tmp_path / "exports"

    path = generate_excel_report(data, replace(branding, logo_path=real_logo), output_dir)

    with zipfile.ZipFile(path) as archive:
        drawing_xml = archive.read("xl/drawings/drawing1.xml")
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}
    ext = ET.fromstring(drawing_xml).find(".//a:ext", ns)
    saved_ratio = float(ext.get("cx")) / float(ext.get("cy"))

    assert saved_ratio == pytest.approx(2.12, abs=0.05)  # não 1.0 (quadrado)


def test_generate_word_report_creates_file(populated_session_id, branding, tmp_path: Path) -> None:
    db, session_id = populated_session_id
    data = _data(db, session_id)
    output_dir = tmp_path / "exports"

    path = generate_word_report(data, branding, output_dir)

    assert path.exists()
    assert path.suffix == ".docx"
    assert path.stat().st_size > 0


def test_generate_pdf_report_creates_file(populated_session_id, branding, tmp_path: Path) -> None:
    db, session_id = populated_session_id
    data = _data(db, session_id)
    output_dir = tmp_path / "exports"

    path = generate_pdf_report(data, branding, output_dir)

    assert path.exists()
    assert path.suffix == ".pdf"
    assert path.stat().st_size > 0


def test_generate_report_honors_operator_chosen_name(populated_session_id, branding, tmp_path: Path) -> None:
    """Nome digitado no diálogo 'Salvar como' do Windows é respeitado (sem extensão)."""
    db, session_id = populated_session_id
    data = _data(db, session_id)
    output_dir = tmp_path / "exports"

    path = generate_word_report(data, branding, output_dir, base_name="Relatorio Final")

    assert path.name == "Relatorio Final.docx"
    assert path.exists()


def test_assemble_report_data_computes_step_stats(populated_session_id) -> None:
    db, session_id = populated_session_id
    data = _data(db, session_id)

    assert len(data.step_stats) == 1  # passo único na fixture
    st = data.step_stats[0]
    assert st.step_index == 0
    assert st.sample_count == 5
    assert st.voltage_min == pytest.approx(5.0)
    assert st.voltage_max == pytest.approx(5.04)
    # Todas as amostras estão dentro da faixa 4.5–5.5 e abaixo de 1.0 A.
    assert st.voltage_out_of_range == 0
    assert st.current_over_limit == 0


def test_traceability_identity_persists_and_reaches_context(populated_session_id, branding) -> None:
    from reports.template_engine import build_context

    db, session_id = populated_session_id
    idn = "Agilent Technologies,E3633A,MY12345678,1.6-5.0-2.0"
    TestSessionRepository(db).set_instrument_identity(session_id, idn)

    data = _data(db, session_id)
    assert data.session.instrument_identity == idn

    context = build_context(data, branding)
    assert context["instrument_identity"] == idn


def test_traceability_calibration_frozen_in_snapshot(populated_session_id, branding) -> None:
    """Modelo/patrimônio/calibração vêm do config_snapshot congelado da sessão."""
    from reports.template_engine import build_context

    db, session_id = populated_session_id
    # Simula o que o main_window grava no snapshot no início do ensaio.
    db.connection.execute(
        "UPDATE test_sessions SET config_snapshot_json = ?, app_version = ? WHERE id = ?",
        (
            '{"voltage_min": 4.5, "voltage_max": 5.5, "current_max": 1.0, '
            '"instrument_model": "Keysight E3633A", "instrument_asset_id": "AVB-0042", '
            '"instrument_calibration_due": "2026-12-31"}',
            "0.1.0",
            session_id,
        ),
    )
    db.connection.commit()

    context = build_context(_data(db, session_id), branding)
    assert context["instrument_model"] == "Keysight E3633A"
    assert context["instrument_asset_id"] == "AVB-0042"
    assert context["instrument_calibration_due"] == "2026-12-31"
    assert context["app_version"] == "0.1.0"


def test_format_duration_human() -> None:
    from reports.template_engine import format_duration

    assert format_duration(45) == "45 s"
    assert format_duration(90) == "1 min 30 s"
    assert format_duration(300) == "5 min"
    assert format_duration(3600) == "1 h"
    assert format_duration(5400) == "1 h 30 min"
    assert format_duration(None) == "—"


def test_render_samples_chart_creates_png(populated_session_id, branding, tmp_path: Path) -> None:
    from reports.chart import render_samples_chart

    db, session_id = populated_session_id
    data = _data(db, session_id)
    out = render_samples_chart(data, branding, tmp_path / "chart.png")

    assert out is not None and out.exists()
    assert out.read_bytes()[:8] == b"\x89PNG\r\n\x1a\n"  # assinatura PNG


def test_excel_report_has_separate_sheets_with_numeric_samples(
    populated_session_id, branding, tmp_path: Path
) -> None:
    from openpyxl import load_workbook

    db, session_id = populated_session_id
    data = _data(db, session_id)
    path = generate_excel_report(data, branding, tmp_path / "exports")

    wb = load_workbook(path)
    # Aba "Gráfico" foi adicionada entre Amostras e Eventos (gráfico dual-eixo dedicado).
    assert wb.sheetnames == ["Resumo", "Amostras", "Gráfico", "Eventos"]

    ws = wb["Amostras"]
    # 1 cabeçalho + 5 amostras.
    assert ws.max_row == 6
    # Tensão/corrente gravadas como número real (não texto).
    assert isinstance(ws["C2"].value, (int, float))
    assert isinstance(ws["D2"].value, (int, float))
    # Coluna E = Tempo(s) como número real, sempre >= 0.
    assert isinstance(ws["E2"].value, (int, float))
    assert ws["E2"].value == 0.0  # primeira amostra: t=0
    assert ws["E6"].value >= 0.0  # última amostra: t > 0


def test_generate_reports_without_evaluation_yet(populated_session_id, branding, tmp_path: Path) -> None:
    """Sessão ainda sem avaliação manual: relatório deve mostrar 'Pendente', sem quebrar."""
    db, session_id = populated_session_id
    db.connection.execute("DELETE FROM evaluations WHERE test_session_id = ?", (session_id,))
    db.connection.commit()
    data = _data(db, session_id)

    assert data.evaluation is None
    path = generate_excel_report(data, branding, tmp_path / "exports")
    assert path.exists()
