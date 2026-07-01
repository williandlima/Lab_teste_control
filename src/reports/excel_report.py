"""Relatório FCT em Excel (.xlsx) via openpyxl, layout vindo de `report_template.yaml`.

Diferente do Word/PDF (documentos de leitura), a planilha é uma FERRAMENTA DE
ANÁLISE. Por isso:

- as amostras vão COMPLETAS (sem subamostrar) numa aba própria e como NÚMEROS
  reais (não texto), com timestamps e tempo decorrido em segundos — filtrável;
- coluna "Tempo (s)" permite plotagem direta no Excel, sem manipular datas;
- formatação condicional realça em vermelho as leituras fora da faixa de
  referência (tensão min/máx; corrente acima do limite) — informativo, nunca
  veredito automático;
- gráfico dedicado (aba "Gráfico") com dois eixos Y independentes, idêntico em
  hierarquia visual ao gráfico do modo ensaio: corrente (teal, linha grossa) no
  eixo direito, tensão (laranja, linha fina) no eixo esquerdo — corrente é a
  grandeza primária do ensaio, tensão é preset do procedimento;
- abas separadas: Resumo, Amostras, Gráfico, Eventos.
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as XlImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import BrandingConfig
from reports.report_data import ReportData
from reports.template_engine import (
    build_context,
    load_template,
    render_fields,
    resolve_output_path,
    result_color_hex,
    step_stats_rows,
)

_LABEL_COL_WIDTH = 32
_VALUE_COL_WIDTH = 48

# Cores idênticas ao LiveChart (consistência ensaio → relatório)
_HEX_VOLTAGE = "FF7A29"   # laranja
_HEX_CURRENT = "1F9E91"   # teal
_HEX_LIMIT = "E74C3C"     # vermelho — mesma cor das linhas-guia do LiveChart
_LINE_W_VOLTAGE = 15875   # ~1.25 pt em EMUs (12700 EMUs = 1 pt)
_LINE_W_CURRENT = 31750   # ~2.5 pt — corrente é a grandeza primária


def _hex_fill(hex_color: str) -> PatternFill:
    h = hex_color.lstrip("#")
    return PatternFill(start_color=h, end_color=h, fill_type="solid")


def _configure_print(ws: Worksheet, landscape: bool = False, repeat_header_row: bool = False) -> None:
    """Ajusta a página para impressão: cabe na largura, sem colunas cortadas."""
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if repeat_header_row:
        ws.print_title_rows = "1:1"


def _voltage_reference_limits(data: ReportData) -> tuple[float, float] | None:
    """(v_min, v_max) configurados para as linhas-guia do gráfico, ou None se ausentes.

    Fonte única de verdade: `_build_samples_sheet` (colunas F/G) e
    `_build_chart_sheet` (linhas-guia tracejadas) chamam esta função — sem
    isso, cada uma reavaliava a mesma condição de forma independente, e uma
    diferença futura entre elas faria o gráfico referenciar colunas nunca
    escritas (ou vice-versa) sem erro nenhum.
    """
    cfg = data.config_snapshot
    v_lo = cfg.get("voltage_min")
    v_hi = cfg.get("voltage_max")
    if v_lo is None or v_hi is None:
        return None
    return v_lo, v_hi


def _write_heading(ws: Worksheet, row: int, text: str, branding: BrandingConfig, span: int = 2) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color=branding.color_text_on_navy.lstrip("#"), size=12)
    cell.fill = _hex_fill(branding.color_secondary_navy)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def _write_fields(ws: Worksheet, row: int, fields: list[tuple[str, str]]) -> int:
    for label, value in fields:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row + 1


def _write_table(ws: Worksheet, row: int, columns: list[str], rows: list[list], branding: BrandingConfig) -> int:
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color=branding.color_text_on_navy.lstrip("#"))
        cell.fill = _hex_fill(branding.color_primary_navy)
    row += 1
    for data_row in rows:
        for col_idx, value in enumerate(data_row, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    return row + 1


def _parse_timestamp(ts: str | None):
    """Devolve datetime real quando possível (para o Excel tratar como data)."""
    if not ts:
        return ts
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt.datetime.strptime(ts, fmt)
        except (ValueError, TypeError):
            pass
    return ts


def _build_summary_sheet(ws: Worksheet, data: ReportData, branding: BrandingConfig, template: dict, context: dict) -> None:
    ws.column_dimensions["A"].width = _LABEL_COL_WIDTH
    ws.column_dimensions["B"].width = _VALUE_COL_WIDTH

    row = 1
    if branding.logo_path.exists():
        img = XlImage(str(branding.logo_path))
        img.height = 60
        img.width = 60
        ws.add_image(img, "A1")
        row = 5

    title_cell = ws.cell(row=row, column=1, value=template["title"])
    title_cell.font = Font(bold=True, size=16, color=branding.color_primary_navy.lstrip("#"))
    row += 1
    ws.cell(row=row, column=1, value=context["company_name"]).font = Font(italic=True, size=11)
    row += 2

    for key in ("identification", "parameters", "execution", "traceability"):
        section = template["sections"][key]
        row = _write_heading(ws, row, section["heading"], branding)
        row = _write_fields(ws, row, render_fields(section["fields"], context))

    evaluation_section = template["sections"]["evaluation"]
    row = _write_heading(ws, row, evaluation_section["heading"], branding)
    eval_fields = render_fields(evaluation_section["fields"], context)
    result_row = row
    row = _write_fields(ws, row, eval_fields)
    color = result_color_hex(data, template, branding)
    if color:
        ws.cell(row=result_row, column=2).fill = _hex_fill(color)

    stats_section = template["sections"]["step_stats_table"]
    stats_rows = step_stats_rows(data)
    if stats_rows:
        row = _write_heading(ws, row, stats_section["heading"], branding, span=len(stats_section["columns"]))
        # Estatísticas como números reais (não as strings formatadas do Word/PDF).
        # Corrente vem antes da tensão nas colunas de média/min/max porque é a
        # grandeza primária do ensaio (tensão é preset do procedimento).
        numeric_rows = [
            [
                st.step_index + 1,
                st.sample_count,
                round(st.current_mean, 4),
                round(st.current_std, 4),
                round(st.current_min, 4),
                round(st.current_max, 4),
                round(st.voltage_mean, 3),
                round(st.voltage_std, 3),
                round(st.voltage_min, 3),
                round(st.voltage_max, 3),
                st.current_over_limit,
                st.voltage_out_of_range,
            ]
            for st in data.step_stats
        ]
        stats_cols = [
            "Ciclo", "Amostras",
            "Corrente méd. (A)", "Corrente desv. (A)", "Corrente mín (A)", "Corrente máx (A)",
            "Tensão méd. (V)", "Tensão desv. (V)", "Tensão mín (V)", "Tensão máx (V)",
            "I acima do limite", "V fora da faixa",
        ]
        _write_table(ws, row, stats_cols, numeric_rows, branding)


def _build_samples_sheet(ws: Worksheet, data: ReportData, branding: BrandingConfig) -> int:
    """Amostras completas como números; retorna a última linha de dados.

    Coluna E ("Tempo (s)") contém o tempo decorrido em segundos desde a
    primeira amostra — permite plotagem direta no Excel sem manipular datas.
    Colunas F/G ("V mín (ref.)"/"V máx (ref.)") repetem os limites configurados
    em toda linha — servem só de fonte para as linhas-guia tracejadas do
    gráfico dual-eixo, replicando o LiveChart do ensaio (seção 11.1).
    """
    limits = _voltage_reference_limits(data)

    columns = ["Timestamp", "Ciclo", "Tensão (V)", "Corrente (A)", "Tempo (s)"]
    if limits is not None:
        columns += ["V mín (ref.)", "V máx (ref.)"]
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=branding.color_text_on_navy.lstrip("#"))
        cell.fill = _hex_fill(branding.color_primary_navy)

    first_dt: dt.datetime | None = None
    for i, s in enumerate(data.samples, start=2):
        ts_val = _parse_timestamp(s.timestamp)
        ts_cell = ws.cell(row=i, column=1, value=ts_val)
        if isinstance(ts_val, dt.datetime):
            ts_cell.number_format = "yyyy-mm-dd hh:mm:ss.000"
            if first_dt is None:
                first_dt = ts_val
            elapsed = round((ts_val - first_dt).total_seconds(), 2)
        else:
            elapsed = float(i - 2)  # fallback: usa índice quando timestamp não parseia
        ws.cell(row=i, column=2, value=s.step_index + 1)
        ws.cell(row=i, column=3, value=round(s.voltage_measured, 4))
        ws.cell(row=i, column=4, value=round(s.current_measured, 4))
        ws.cell(row=i, column=5, value=elapsed)
        if limits is not None:
            ws.cell(row=i, column=6, value=limits[0])
            ws.cell(row=i, column=7, value=limits[1])

    last_row = len(data.samples) + 1
    ws.column_dimensions["A"].width = 24
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["E"].width = 12
    if limits is not None:
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A2"
    if last_row >= 1:
        ws.auto_filter.ref = f"A1:E{max(last_row, 1)}"

    # Formatação condicional: realça leituras fora da faixa de referência.
    if data.samples:
        cfg = data.config_snapshot
        red = _hex_fill(branding.color_fail)
        v_lo, v_hi = cfg.get("voltage_min"), cfg.get("voltage_max")
        i_hi = cfg.get("current_max")
        v_range = f"C2:C{last_row}"
        i_range = f"D2:D{last_row}"
        if v_lo is not None:
            ws.conditional_formatting.add(v_range, CellIsRule(operator="lessThan", formula=[str(v_lo)], fill=red))
        if v_hi is not None:
            ws.conditional_formatting.add(v_range, CellIsRule(operator="greaterThan", formula=[str(v_hi)], fill=red))
        if i_hi is not None:
            ws.conditional_formatting.add(i_range, CellIsRule(operator="greaterThan", formula=[str(i_hi)], fill=red))
    return last_row


def _build_chart_sheet(ws_chart: Worksheet, ws_samples: Worksheet, last_row: int, data: ReportData) -> None:
    """Gráfico dual-eixo: corrente (primária, direita, teal grossa) × tensão (esquerda, laranja fina).

    O eixo X usa a coluna "Tempo (s)" (numérica) — elimina o problema dos
    rótulos de timestamp que apareciam ilegíveis no gráfico anterior.
    A hierarquia visual reflete a hierarquia do ensaio: corrente é o que
    o operador veio observar; tensão é o preset que viabiliza a observação.
    """
    if last_row < 2:
        return

    limits = _voltage_reference_limits(data)
    i_hi_ref = data.config_snapshot.get("current_max")

    # -- Chart primário: Tensão no eixo Y esquerdo --
    c_voltage = LineChart()
    c_voltage.title = "Corrente e Tensão × Tempo (ensaio)"
    c_voltage.height = 15
    c_voltage.width = 26
    c_voltage.y_axis.title = "Tensão (V)"
    c_voltage.y_axis.axId = 100
    c_voltage.y_axis.axPos = "l"
    c_voltage.y_axis.numFmt = '0.00"V"'
    c_voltage.x_axis.axId = 10
    # openpyxl cria todo eixo com axPos="l" por padrão (mesmo o de categorias) —
    # sem esta linha o eixo do tempo fica desenhado na vertical à esquerda,
    # sobrepondo o eixo de tensão, e o Excel real (diferente do LibreOffice)
    # recusa a plotar as séries: gráfico aparece em branco/sem valores.
    c_voltage.x_axis.axPos = "b"
    c_voltage.x_axis.title = "Tempo (s)"
    c_voltage.x_axis.numFmt = '0"s"'

    if limits is not None:
        v_lo_ref, v_hi_ref = limits
        margin = max(0.5, (v_hi_ref - v_lo_ref) * 0.2)
        c_voltage.y_axis.scaling.min = round(v_lo_ref - margin, 2)
        c_voltage.y_axis.scaling.max = round(v_hi_ref + margin, 2)

    v_ref = Reference(ws_samples, min_col=3, max_col=3, min_row=1, max_row=last_row)
    c_voltage.add_data(v_ref, titles_from_data=True)
    # Tensão: laranja, linha fina (informação de preset)
    c_voltage.series[0].graphicalProperties.line.solidFill = _HEX_VOLTAGE
    c_voltage.series[0].graphicalProperties.line.width = _LINE_W_VOLTAGE

    # Linhas-guia tracejadas de V mín/máx (colunas F/G) — mesma referência
    # visual do LiveChart do ensaio (linha vermelha tracejada), nunca um
    # veredito automático.
    if limits is not None:
        limit_ref = Reference(ws_samples, min_col=6, max_col=7, min_row=1, max_row=last_row)
        series_before = len(c_voltage.series)
        c_voltage.add_data(limit_ref, titles_from_data=True)
        # Captura pelas séries recém-adicionadas (não por índice fixo): robusto
        # a qualquer add_data() futuro inserido antes deste bloco.
        for limit_series in c_voltage.series[series_before:]:
            limit_series.graphicalProperties.line.solidFill = _HEX_LIMIT
            limit_series.graphicalProperties.line.width = 12700  # 1 pt
            limit_series.graphicalProperties.line.dashStyle = "dash"

    # -- Chart secundário: Corrente no eixo Y direito (grandeza primária) --
    c_current = LineChart()
    c_current.y_axis.title = "Corrente (A)"
    c_current.y_axis.axId = 200
    c_current.y_axis.axPos = "r"
    c_current.y_axis.crossAx = 100
    c_current.y_axis.crosses = "max"   # eixo Y da corrente fica na direita
    c_current.y_axis.numFmt = '0.000"A"'
    c_current.x_axis.axId = 10
    c_current.x_axis.delete = True    # não duplicar o eixo X

    if i_hi_ref is not None:
        c_current.y_axis.scaling.min = 0
        c_current.y_axis.scaling.max = round(max(i_hi_ref * 1.2, 0.01), 4)

    i_ref = Reference(ws_samples, min_col=4, max_col=4, min_row=1, max_row=last_row)
    c_current.add_data(i_ref, titles_from_data=True)
    # Corrente: teal, linha grossa (grandeza primária do ensaio)
    c_current.series[0].graphicalProperties.line.solidFill = _HEX_CURRENT
    c_current.series[0].graphicalProperties.line.width = _LINE_W_CURRENT

    # Categorias: coluna "Tempo (s)" (col E = índice 5) — numéricas, não timestamps
    time_ref = Reference(ws_samples, min_col=5, min_row=2, max_row=last_row)
    c_voltage.set_categories(time_ref)

    # Mesclar os dois charts (voltage recebe current como eixo secundário)
    c_voltage += c_current

    ws_chart.add_chart(c_voltage, "A1")


def generate_excel_report(
    data: ReportData, branding: BrandingConfig, output_dir: Path, base_name: str | None = None
) -> Path:
    template = load_template()
    context = build_context(data, branding)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumo"
    ws_samples = wb.create_sheet("Amostras")
    ws_chart = wb.create_sheet("Gráfico")
    ws_events = wb.create_sheet("Eventos")
    # Aba do gráfico em teal — mesma cor da corrente, a grandeza primária.
    ws_chart.sheet_properties.tabColor = _HEX_CURRENT

    _build_summary_sheet(ws_summary, data, branding, template, context)
    last_row = _build_samples_sheet(ws_samples, data, branding)
    _build_chart_sheet(ws_chart, ws_samples, last_row, data)

    events_columns = template["sections"]["events_table"]["columns"]
    event_rows = [[e.timestamp or "", e.level, e.source, e.message] for e in data.events]
    _write_table(ws_events, 1, events_columns, event_rows, branding)
    ws_events.column_dimensions["A"].width = 24
    ws_events.column_dimensions["D"].width = 60
    ws_events.freeze_panes = "A2"

    _configure_print(ws_summary, landscape=False, repeat_header_row=False)
    _configure_print(ws_samples, landscape=True, repeat_header_row=True)
    _configure_print(ws_chart, landscape=True, repeat_header_row=False)
    _configure_print(ws_events, landscape=True, repeat_header_row=True)

    footer_cell = ws_summary.cell(row=ws_summary.max_row + 2, column=1, value=template["footer"].format(**context))
    footer_cell.font = Font(italic=True, size=9)
    footer_cell.alignment = Alignment(horizontal="left")

    output_path = resolve_output_path(data, output_dir, "xlsx", base_name)
    wb.save(output_path)
    return output_path
