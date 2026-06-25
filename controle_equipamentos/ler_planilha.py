"""Leitura (somente leitura) da planilha principal de cadastro de equipamentos."""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .config import AppConfig
from .io_utils import copiar_para_temp
from .modelos import Equipamento


def _valor_texto(valor: object) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _valor_data(valor: object) -> date | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _indice_colunas(cabecalho: list[object], nomes_esperados: dict[str, str | None]) -> dict[str, int | None]:
    """Mapeia cada chave lógica (ex.: 'codigo') para o índice da coluna na planilha."""
    cabecalho_normalizado = [_valor_texto(c).casefold() for c in cabecalho]
    indices: dict[str, int | None] = {}
    for chave, nome_coluna in nomes_esperados.items():
        if nome_coluna is None:
            indices[chave] = None
            continue
        try:
            indices[chave] = cabecalho_normalizado.index(nome_coluna.casefold())
        except ValueError:
            indices[chave] = None
    return indices


def ler_equipamentos(config: AppConfig) -> list[Equipamento]:
    """Lê a planilha de equipamentos e devolve a lista de cadastros.

    Se a coluna de "Próxima Calibração" não existir na planilha (config ou
    cabeçalho real), `proxima_calibracao` vem None para todos os registros —
    quem calcula a faixa de alerta decide o fallback (ver classificar.py).
    """
    caminho = copiar_para_temp(config.planilhas.equipamentos_path, config.planilhas.pasta_temp)
    try:
        workbook = load_workbook(caminho, data_only=True, read_only=True)
        sheet = (
            workbook[config.planilhas.equipamentos_sheet]
            if config.planilhas.equipamentos_sheet
            else workbook.worksheets[0]
        )

        linhas = sheet.iter_rows(values_only=True)
        cabecalho = list(next(linhas))

        cols = config.colunas_equipamentos
        indices = _indice_colunas(
            cabecalho,
            {
                "codigo": cols.codigo,
                "descricao": cols.descricao,
                "responsavel_padrao": cols.responsavel_padrao,
                "local_padrao": cols.local_padrao,
                "status": cols.status,
                "proxima_calibracao": cols.proxima_calibracao,
                "email_responsavel": cols.email_responsavel,
            },
        )

        def campo(linha: tuple, chave: str) -> object:
            indice = indices[chave]
            if indice is None or indice >= len(linha):
                return None
            return linha[indice]

        equipamentos: list[Equipamento] = []
        for linha in linhas:
            codigo = _valor_texto(campo(linha, "codigo"))
            if not codigo:
                continue
            equipamentos.append(
                Equipamento(
                    codigo=codigo,
                    descricao=_valor_texto(campo(linha, "descricao")),
                    responsavel_padrao=_valor_texto(campo(linha, "responsavel_padrao")),
                    local_padrao=_valor_texto(campo(linha, "local_padrao")),
                    status=_valor_texto(campo(linha, "status")),
                    proxima_calibracao=_valor_data(campo(linha, "proxima_calibracao")),
                    email_responsavel=_valor_texto(campo(linha, "email_responsavel")) or None,
                )
            )
        return equipamentos
    finally:
        caminho.unlink(missing_ok=True)
