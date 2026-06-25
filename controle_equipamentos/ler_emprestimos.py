"""Leitura (somente leitura) do log de empréstimos (emprestimos.xlsx).

O arquivo é só de inserção: cada linha é um evento (Retirada/Devolução) e
nunca é editado depois de gravado. Este módulo só lê e ordena por data/hora —
quem calcula "quem está com o quê agora" é localizacao_atual.py.
"""
from __future__ import annotations

from datetime import date, datetime

from openpyxl import load_workbook

from .config import AppConfig
from .io_utils import copiar_para_temp
from .modelos import EventoEmprestimo, TipoEvento


def _valor_texto(valor: object) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _valor_data(valor: object) -> date | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _valor_data_hora(valor: object) -> datetime | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)
    texto = str(valor).strip()
    for formato in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(texto, formato)
        except ValueError:
            continue
    return None


def _tipo_evento(valor: object) -> TipoEvento | None:
    texto = _valor_texto(valor).casefold()
    if texto.startswith("retirada"):
        return TipoEvento.RETIRADA
    if texto.startswith("devolu"):
        return TipoEvento.DEVOLUCAO
    return None


def _indice_colunas(cabecalho: list[object], nomes_esperados: dict[str, str]) -> dict[str, int | None]:
    cabecalho_normalizado = [_valor_texto(c).casefold() for c in cabecalho]
    indices: dict[str, int | None] = {}
    for chave, nome_coluna in nomes_esperados.items():
        try:
            indices[chave] = cabecalho_normalizado.index(nome_coluna.casefold())
        except ValueError:
            indices[chave] = None
    return indices


def ler_emprestimos(config: AppConfig) -> list[EventoEmprestimo]:
    """Lê emprestimos.xlsx e devolve os eventos ordenados por data/hora.

    Linhas sem código de equipamento, sem data/hora válida, ou com tipo de
    evento não reconhecido (diferente de Retirada/Devolução) são ignoradas
    silenciosamente — provavelmente é uma linha em branco no final.
    """
    caminho = copiar_para_temp(config.planilhas.emprestimos_path, config.planilhas.pasta_temp)
    try:
        workbook = load_workbook(caminho, data_only=True, read_only=True)
        sheet = (
            workbook[config.planilhas.emprestimos_sheet]
            if config.planilhas.emprestimos_sheet
            else workbook.worksheets[0]
        )

        linhas = sheet.iter_rows(values_only=True)
        cabecalho = list(next(linhas))

        cols = config.colunas_emprestimos
        indices = _indice_colunas(
            cabecalho,
            {
                "data_hora": cols.data_hora,
                "codigo": cols.codigo,
                "tipo_evento": cols.tipo_evento,
                "responsavel": cols.responsavel,
                "destino": cols.destino,
                "previsao_devolucao": cols.previsao_devolucao,
                "observacao": cols.observacao,
            },
        )

        def campo(linha: tuple, chave: str) -> object:
            indice = indices[chave]
            if indice is None or indice >= len(linha):
                return None
            return linha[indice]

        eventos: list[EventoEmprestimo] = []
        for linha in linhas:
            codigo = _valor_texto(campo(linha, "codigo"))
            data_hora = _valor_data_hora(campo(linha, "data_hora"))
            tipo_evento = _tipo_evento(campo(linha, "tipo_evento"))
            if not codigo or data_hora is None or tipo_evento is None:
                continue
            eventos.append(
                EventoEmprestimo(
                    data_hora=data_hora,
                    codigo=codigo,
                    tipo_evento=tipo_evento,
                    responsavel=_valor_texto(campo(linha, "responsavel")),
                    destino=_valor_texto(campo(linha, "destino")),
                    previsao_devolucao=_valor_data(campo(linha, "previsao_devolucao")),
                    observacao=_valor_texto(campo(linha, "observacao")),
                )
            )

        eventos.sort(key=lambda evento: evento.data_hora)
        return eventos
    finally:
        caminho.unlink(missing_ok=True)
