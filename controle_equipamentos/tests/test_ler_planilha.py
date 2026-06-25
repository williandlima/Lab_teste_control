from datetime import date
from pathlib import Path

from openpyxl import Workbook

from controle_equipamentos.config import AppConfig, ColunasEquipamentosConfig, PlanilhasConfig
from controle_equipamentos.ler_planilha import ler_equipamentos


def _config(equipamentos_path: Path, pasta_temp: Path) -> AppConfig:
    return AppConfig(
        planilhas=PlanilhasConfig(
            equipamentos_path=equipamentos_path,
            equipamentos_sheet=None,
            emprestimos_path=equipamentos_path,
            emprestimos_sheet=None,
            pasta_temp=pasta_temp,
        ),
        colunas_equipamentos=ColunasEquipamentosConfig(
            codigo="Código",
            descricao="Descrição",
            responsavel_padrao="Responsável",
            local_padrao="Local",
            status="Status",
            proxima_calibracao="Próxima Calibração",
            email_responsavel="E-mail",
        ),
        colunas_emprestimos=None,  # type: ignore[arg-type]
        status_ignorados=(),
        limiares_dias=None,  # type: ignore[arg-type]
        smtp=None,  # type: ignore[arg-type]
        estado=None,  # type: ignore[arg-type]
        raw={},
    )


def test_le_equipamentos_com_data_de_calibracao(tmp_path: Path):
    caminho = tmp_path / "Equipamentos.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Código", "Descrição", "Responsável", "Local", "Status", "Próxima Calibração", "E-mail"])
    sheet.append(["A1", "Multímetro", "Setor X", "Sala 1", "Em uso", date(2026, 12, 1), "x@empresa.com"])
    workbook.save(caminho)

    equipamentos = ler_equipamentos(_config(caminho, tmp_path / "tmp"))

    assert len(equipamentos) == 1
    equipamento = equipamentos[0]
    assert equipamento.codigo == "A1"
    assert equipamento.proxima_calibracao == date(2026, 12, 1)
    assert equipamento.email_responsavel == "x@empresa.com"


def test_linha_sem_codigo_e_ignorada(tmp_path: Path):
    caminho = tmp_path / "Equipamentos.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Código", "Descrição", "Responsável", "Local", "Status", "Próxima Calibração", "E-mail"])
    sheet.append([None, None, None, None, None, None, None])
    workbook.save(caminho)

    equipamentos = ler_equipamentos(_config(caminho, tmp_path / "tmp"))

    assert equipamentos == []


def test_coluna_de_calibracao_ausente_fica_none(tmp_path: Path):
    caminho = tmp_path / "Equipamentos.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Código", "Descrição", "Responsável", "Local", "Status"])
    sheet.append(["A1", "Multímetro", "Setor X", "Sala 1", "Aguardando Calibração"])
    workbook.save(caminho)

    config = _config(caminho, tmp_path / "tmp")
    equipamentos = ler_equipamentos(config)

    assert equipamentos[0].proxima_calibracao is None
